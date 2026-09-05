from __future__ import annotations

import re
import io
import math
import threading
import traceback
import zipfile
from contextlib import ExitStack
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tkinter import X, filedialog, messagebox, ttk
import tkinter as tk


LINE_PATTERN = re.compile(
    r"^(?P<date>\d{2}\.\d{2}\.\d{2})\s+"
    r"(?P<time>\d{2}:\d{2}:\d{2})\s+"
    r"(?P<first>\d+)\s+000\s+(?P<second>\d+)\s*$"
)


@dataclass(frozen=True)
class Measurement:
    timestamp: datetime
    first_speed: int
    second_speed: int
    source_file: str
    vehicle_number: int


def parse_text_lines(lines, source_file: str, vehicle_start: int, progress=None) -> tuple[list[Measurement], int, int]:
    measurements: list[Measurement] = []
    pending: tuple[datetime, int, int] | None = None
    vehicle_number = vehicle_start
    line_count = 0
    for line_number, raw_line in enumerate(lines, 1):
        line_count += 1
        line = raw_line.strip()
        if not line:
            continue
        match = LINE_PATTERN.match(line)
        if not match:
            raise ValueError(f"{source_file}, line {line_number}: unexpected format")
        try:
            day, month, year = (int(value) for value in match["date"].split("."))
            hour, minute, second = (int(value) for value in match["time"].split(":"))
            timestamp = datetime(2000 + year, month, day, hour, minute, second)
            current = (timestamp, int(match["first"]), int(match["second"]))
            if pending is None:
                pending = current
            else:
                measurements.extend([
                    Measurement(pending[0], pending[1], pending[2], source_file, vehicle_number),
                    Measurement(current[0], current[1], current[2], source_file, vehicle_number),
                ])
                vehicle_number += 1
                pending = None
        except ValueError as error:
            raise ValueError(f"{source_file}, line {line_number}: {error}") from error

        if progress:
            progress(line_count)
    return measurements, vehicle_number, line_count


def parse_text_file(text: str, source_file: str, vehicle_start: int) -> tuple[list[Measurement], int]:
    measurements, vehicle_number, _ = parse_text_lines(text.splitlines(), source_file, vehicle_start)
    return measurements, vehicle_number


def load_archives(paths: list[str], progress=None) -> tuple[list[Measurement], int]:
    measurements: list[Measurement] = []
    skipped_lines = 0
    vehicle_number = 1
    text_sources: list[tuple[str, str, str | None]] = []
    for archive_path in paths:
        if archive_path.lower().endswith(".txt"):
            text_sources.append((Path(archive_path).name, archive_path, None))
        else:
            with zipfile.ZipFile(archive_path) as archive:
                text_sources.extend(
                    (f"{Path(archive_path).name}/{entry.filename}", archive_path, entry.filename)
                    for entry in archive.infolist()
                    if not entry.is_dir() and entry.filename.lower().endswith(".txt")
                )
    total_sources = len(text_sources)
    archive_paths = {archive_path for _, archive_path, entry_name in text_sources if entry_name is not None}
    with ExitStack() as stack:
        archives = {archive_path: stack.enter_context(zipfile.ZipFile(archive_path)) for archive_path in archive_paths}
        for source_index, (source_name, source_path, entry_name) in enumerate(text_sources, 1):
            if entry_name is not None:
                raw_file = archives[source_path].open(entry_name)
                text_file = io.TextIOWrapper(raw_file, encoding="utf-8-sig", errors="replace")
                try:
                    parsed, vehicle_number, line_count = parse_text_lines(text_file, source_name, vehicle_number)
                finally:
                    text_file.close()
            else:
                with open(source_path, "r", encoding="utf-8-sig", errors="replace") as text_file:
                    parsed, vehicle_number, line_count = parse_text_lines(text_file, source_name, vehicle_number)
            skipped_lines += line_count - len(parsed)
            measurements.extend(parsed)
            if progress:
                progress(source_index, total_sources, source_name)
    measurements.sort(key=lambda item: item.timestamp)
    return measurements, skipped_lines


def export_workbook(measurements: list[Measurement], output_path: str, progress=None) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.drawing.colors import ColorChoice
        from openpyxl.styles import PatternFill
    except ImportError as error:
        raise RuntimeError("The openpyxl package is required to create Excel files.") from error

    workbook = Workbook()
    report = workbook.active
    report.title = "Report"
    report.print_title_rows = "1:1"
    report.sheet_properties.pageSetUpPr.fitToPage = True
    report.page_setup.fitToWidth = 1
    report.page_setup.fitToHeight = 0
    headers = ["Date", "Time", "Vehicle", "First speed (km/h)",
               "Second speed (km/h)", "Reduction (km/h)"]
    summary_col = 1
    daily_col = 1
    hourly_col = 1
    chart_data_col = 1
    chart_x_col = 8
    summary_start = 150

    def timestamp_text(value: datetime) -> str:
        return value.strftime("%d.%m.%Y %H:%M:%S")

    def excel_serial(value: datetime) -> float:
        return (value - datetime(1899, 12, 30)).total_seconds() / 86400

    def write_cell(row: int, column: int, value) -> None:
        report.cell(row, column, value)

    def stats(items: list[Measurement]) -> tuple[Measurement, Measurement]:
        minimum = min(items, key=lambda item: min(item.first_speed, item.second_speed))
        maximum = max(items, key=lambda item: max(item.first_speed, item.second_speed))
        return minimum, maximum

    minimum, maximum = stats(measurements)
    write_cell(summary_start, summary_col, "Selected data report")
    write_cell(summary_start + 1, summary_col, "Only dates selected in the date dialog are included below and in all charts.")
    write_cell(summary_start + 3, summary_col, "Overall statistics")
    overall = [
        ("Total vehicles", len(measurements) // 2),
        ("Total measurements", len(measurements)),
        ("Minimum speed (km/h)", min(min(item.first_speed, item.second_speed) for item in measurements)),
        ("Minimum recorded at", timestamp_text(minimum.timestamp)),
        ("Maximum speed (km/h)", max(max(item.first_speed, item.second_speed) for item in measurements)),
        ("Maximum recorded at", timestamp_text(maximum.timestamp)),
    ]
    for row, (label, value) in enumerate(overall, summary_start + 4):
        write_cell(row, summary_col, label)
        write_cell(row, summary_col + 1, value)
    by_date: dict[date, list[Measurement]] = defaultdict(list)
    by_hour: dict[int, list[Measurement]] = defaultdict(list)
    for item in measurements:
        by_date[item.timestamp.date()].append(item)
        by_hour[item.timestamp.hour].append(item)

    daily_start = summary_start + 12
    write_cell(daily_start, daily_col, "Daily statistics")
    daily_headers = ["Date", "Vehicles", "Minimum speed", "Minimum recorded at", "Maximum speed", "Maximum recorded at", "Average reduction"]
    for offset, value in enumerate(daily_headers):
        write_cell(daily_start + 1, daily_col + offset, value)
    for row, current_date in enumerate(sorted(by_date), daily_start + 2):
        items = by_date[current_date]
        day_min, day_max = stats(items)
        values = [current_date.strftime("%d.%m.%Y"), len(items) // 2, min(min(i.first_speed, i.second_speed) for i in items),
              timestamp_text(day_min.timestamp), max(max(i.first_speed, i.second_speed) for i in items), timestamp_text(day_max.timestamp),
                  sum(i.first_speed - i.second_speed for i in items) / len(items)]
        for offset, value in enumerate(values):
            write_cell(row, daily_col + offset, value)

    hourly_start = daily_start + 3 + len(by_date)
    write_cell(hourly_start, hourly_col, "Hourly statistics across selected dates")
    hourly_headers = ["Hour", "Measurements", "Minimum speed", "Minimum recorded at", "Maximum speed", "Maximum recorded at", "Average reduction"]
    for offset, value in enumerate(hourly_headers):
        write_cell(hourly_start + 1, hourly_col + offset, value)
    for row, hour in enumerate(sorted(by_hour), hourly_start + 2):
        items = by_hour[hour]
        hour_min, hour_max = stats(items)
        values = [f"{hour:02d}:00", len(items), min(min(i.first_speed, i.second_speed) for i in items),
              timestamp_text(hour_min.timestamp), max(max(i.first_speed, i.second_speed) for i in items), timestamp_text(hour_max.timestamp),
                  sum(i.first_speed - i.second_speed for i in items) / len(items)]
        for offset, value in enumerate(values):
            write_cell(row, hourly_col + offset, value)

    chart_headers = ["Date/time", "Minimum speed", "Maximum speed", "Average reduction", "First average speed", "Second average speed"]
    chart_source_start = hourly_start + 3 + len(by_hour)
    write_cell(chart_source_start, chart_data_col, "Chart sources")
    for offset, value in enumerate(chart_headers):
        write_cell(chart_source_start + 1, chart_data_col + offset, value)
    def sampled(values: list, limit: int = 6) -> list:
        if len(values) <= limit:
            return values
        indexes = [round(index * (len(values) - 1) / (limit - 1)) for index in range(limit)]
        return [values[index] for index in indexes]

    chart_dates = sampled(sorted(by_date))
    for row, current_date in enumerate(chart_dates, chart_source_start + 2):
        items = by_date[current_date]
        write_cell(row, chart_data_col, current_date.strftime("%d.%m.%Y"))
        write_cell(row, chart_data_col + 1, min(min(i.first_speed, i.second_speed) for i in items))
        write_cell(row, chart_data_col + 2, max(max(i.first_speed, i.second_speed) for i in items))
        write_cell(row, chart_data_col + 3, sum(i.first_speed - i.second_speed for i in items) / len(items))
        write_cell(row, chart_data_col + 4, sum(i.first_speed for i in items) / len(items))
        write_cell(row, chart_data_col + 5, sum(i.second_speed for i in items) / len(items))
    daily_chart_start = chart_source_start + 1
    daily_chart_end = chart_source_start + 1 + len(chart_dates)
    hourly_chart_start = daily_chart_end + 3
    for offset, value in enumerate(chart_headers):
        write_cell(hourly_chart_start, chart_data_col + offset, value)
    chart_hours = sampled(sorted(by_hour))
    for row, hour in enumerate(chart_hours, hourly_chart_start + 1):
        items = by_hour[hour]
        write_cell(row, chart_data_col, f"{hour:02d}:00")
        write_cell(row, chart_data_col + 1, min(min(i.first_speed, i.second_speed) for i in items))
        write_cell(row, chart_data_col + 2, max(max(i.first_speed, i.second_speed) for i in items))
        write_cell(row, chart_data_col + 3, sum(i.first_speed - i.second_speed for i in items) / len(items))
        write_cell(row, chart_data_col + 4, sum(i.first_speed for i in items) / len(items))
        write_cell(row, chart_data_col + 5, sum(i.second_speed for i in items) / len(items))
    hourly_chart_end = hourly_chart_start + len(chart_hours)

    def add_line_chart(title: str, anchor: str, category_column: int, y_columns: list[int], colors: list[str], start_row: int, end_row: int, x_title: str, y_title: str, y_min: float, y_max: float, major_unit: int) -> None:
        chart = LineChart()
        chart.title = title
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.y_axis.numFmt = "0.0"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.minorTickMark = "none"
        chart.y_axis.minorTickMark = "none"
        chart.y_axis.scaling.min = y_min
        chart.y_axis.scaling.max = y_max
        chart.y_axis.majorUnit = major_unit
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"
        chart.legend = None
        chart.height = 18
        chart.width = 36
        for series_index, y_column in enumerate(y_columns):
            y_values = Reference(report, min_col=y_column, min_row=start_row, max_row=end_row)
            chart.add_data(y_values, titles_from_data=True)
            series = chart.series[-1]
            series.graphicalProperties.line.solidFill = ColorChoice(srgbClr=colors[series_index])
            series.graphicalProperties.line.width = 24000
        categories = Reference(report, min_col=category_column, min_row=start_row + 1, max_row=end_row)
        chart.set_categories(categories)
        report.add_chart(chart, anchor)

    def padded_bounds(values: list[float]) -> tuple[float, float]:
        low = min(values)
        high = max(values)
        padding = max((high - low) * 0.05, 1)
        return low - padding, high + padding

    speed_values = [speed for items in by_date.values() for item in items for speed in (item.first_speed, item.second_speed)]
    reduction_values = [
        sum(item.first_speed - item.second_speed for item in items) / len(items)
        for items in by_date.values()
    ] + [
        sum(item.first_speed - item.second_speed for item in items) / len(items)
        for items in by_hour.values()
    ]
    speed_y_min = 0
    speed_y_max = max(20, math.ceil(max(speed_values) / 20) * 20)
    reduction_y_min = 0
    reduction_y_max = max(5, math.ceil(max(0, max(reduction_values)) / 5) * 5)

    add_line_chart("Daily minimum and maximum speed", "A1", chart_data_col, [chart_data_col + 1, chart_data_col + 2], ["1F77B4", "E67E22"], daily_chart_start, daily_chart_end, "Date", "Speed (km/h)", speed_y_min, speed_y_max, 20)
    add_line_chart("Daily average speed reduction", "A38", chart_data_col, [chart_data_col + 3], ["2CA02C"], daily_chart_start, daily_chart_end, "Date", "Speed reduction (km/h)", reduction_y_min, reduction_y_max, 5)
    add_line_chart("Hourly minimum and maximum speed", "A75", chart_data_col, [chart_data_col + 1, chart_data_col + 2], ["1F77B4", "E67E22"], hourly_chart_start, hourly_chart_end, "Time of day", "Speed (km/h)", speed_y_min, speed_y_max, 20)
    add_line_chart("Hourly average speed reduction", "A112", chart_data_col, [chart_data_col + 3], ["2CA02C"], hourly_chart_start, hourly_chart_end, "Time of day", "Speed reduction (km/h)", reduction_y_min, reduction_y_max, 5)

    write_cell(chart_source_start, chart_x_col, "Numeric chart x-values (Excel serials)")
    for column in range(chart_data_col, chart_data_col + 4):
        report.column_dimensions[report.cell(1, column).column_letter].width = 22
    report.column_dimensions[report.cell(1, chart_x_col).column_letter].hidden = True
    for column in range(1, 8):
        report.column_dimensions[report.cell(1, column).column_letter].width = 24

    first_measurements: dict[int, Measurement] = {}
    for item in measurements:
        previous = first_measurements.get(item.vehicle_number)
        if previous is None or item.timestamp < previous.timestamp:
            first_measurements[item.vehicle_number] = item
    first_vehicle_rows = list(first_measurements.values())
    def configure_heatmap_sheet(sheet) -> None:
        sheet.print_title_rows = "1:2"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

    def write_heatmap(sheet, title: str, columns: list, key_function) -> None:
        counts: dict[tuple[object, int], int] = defaultdict(int)
        totals: dict[object, int] = defaultdict(int)
        max_speed_bin = (max(item.first_speed for item in first_vehicle_rows) // 5) * 5
        for item in first_vehicle_rows:
            key = key_function(item)
            speed_bin = (item.first_speed // 5) * 5
            counts[(key, speed_bin)] += 1
            totals[key] += 1
        sheet.cell(1, 1, title)
        sheet.cell(2, 1, "Speed bin (km/h)")
        for column_index, column_value in enumerate(columns, 2):
            sheet.cell(2, column_index, str(column_value))
            sheet.column_dimensions[sheet.cell(1, column_index).column_letter].width = 12
        for speed_bin in range(0, max_speed_bin + 5, 5):
            row = 3 + speed_bin // 5
            sheet.cell(row, 1, f"{speed_bin}-{speed_bin + 4}")
            for column_index, column_value in enumerate(columns, 2):
                count = counts.get((column_value, speed_bin), 0)
                cell = sheet.cell(row, column_index, count)
                if count:
                    percentage = count / totals[column_value]
                    intensity = int(250 - (percentage ** 0.5) * 235)
                    green = min(255, intensity + 20)
                    cell.fill = PatternFill(fill_type="solid", fgColor=f"FF{intensity:02X}{green:02X}FF")
        sheet.column_dimensions["A"].width = 18

    heatmap_dates = sorted({item.timestamp.date().strftime("%d.%m.%Y") for item in first_vehicle_rows})
    heatmap_hours = sorted({f"{item.timestamp.hour:02d}:00" for item in first_vehicle_rows})
    daily_heatmap = workbook.create_sheet("Speed heatmap by day")
    hourly_heatmap = workbook.create_sheet("Speed heatmap by hour")
    configure_heatmap_sheet(daily_heatmap)
    configure_heatmap_sheet(hourly_heatmap)
    write_heatmap(
        daily_heatmap, "First measurement speed count by day", heatmap_dates,
        lambda item: item.timestamp.date().strftime("%d.%m.%Y"),
    )
    write_heatmap(
        hourly_heatmap, "First measurement speed count by hour", heatmap_hours,
        lambda item: f"{item.timestamp.hour:02d}:00",
    )

    max_data_rows = 1_048_575
    def write_data_sheet(sheet, sheet_measurements: list[Measurement]) -> None:
        sheet.append(headers)
        for item in sheet_measurements:
            sheet.append([item.timestamp.strftime("%d.%m.%Y"), item.timestamp.strftime("%H:%M:%S"), item.vehicle_number,
                          item.first_speed, item.second_speed, item.first_speed - item.second_speed])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:F{len(sheet_measurements) + 1}"
        sheet.print_title_rows = "1:1"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        for column, width in {"A": 13, "B": 11, "C": 10, "D": 19, "E": 20, "F": 17}.items():
            sheet.column_dimensions[column].width = width

    data_chunks = max(1, (len(measurements) + max_data_rows - 1) // max_data_rows)
    if progress:
        progress(15, "Writing measurement data")
    for chunk_number, start in enumerate(range(0, len(measurements), max_data_rows), 1):
        measurement_sheet = workbook.create_sheet(f"Measurements {chunk_number}")
        write_data_sheet(measurement_sheet, measurements[start:start + max_data_rows])
        if progress:
            progress(15 + int(chunk_number / data_chunks * 55), "Writing measurement data")

    if progress:
        progress(90, "Saving Excel workbook")
    workbook.save(output_path)
    if progress:
        progress(100, "Finished")


class ProgressDialog:
    def __init__(self, root: tk.Tk) -> None:
        self.window = tk.Toplevel(root)
        self.window.title("Converting")
        self.window.geometry("520x125")
        self.window.resizable(False, False)
        self.window.transient(root)
        self.window.grab_set()
        self.message = tk.StringVar(value="Preparing...")
        ttk.Label(self.window, textvariable=self.message, padding=(16, 14, 16, 8)).pack(fill=X)
        self.progress = ttk.Progressbar(self.window, mode="determinate", maximum=100)
        self.progress.pack(fill=X, padx=16, pady=(0, 16))

    def update(self, percent: int, message: str) -> None:
        self.progress["value"] = percent
        self.message.set(f"{message} ({percent}%)")


class DateSelectionDialog:
    def __init__(self, root: tk.Tk, dates: list[date]) -> None:
        self.selected: list[date] | None = None
        self.window = tk.Toplevel(root)
        self.window.title("Select dates")
        self.window.geometry("360x480")
        self.window.transient(root)
        self.window.grab_set()
        ttk.Label(self.window, text="Select the dates to include in the Excel report.\nKeep pressing SHIFT or CTRL while clicking\nto select multiple dates.", padding=12).pack(anchor="w")
        body = ttk.Frame(self.window)
        body.pack(fill="both", expand=True, padx=12)
        self.date_list = tk.Listbox(body, selectmode=tk.EXTENDED, exportselection=False,
                                    activestyle="none", height=20)
        scrollbar = ttk.Scrollbar(body, orient="vertical", command=self.date_list.yview)
        self.date_list.configure(yscrollcommand=scrollbar.set)
        self.date_list.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.dates = dates
        for current_date in dates:
            self.date_list.insert(tk.END, current_date.strftime("%d.%m.%Y"))
        if dates:
            self.date_list.selection_set(0, tk.END)
        buttons = ttk.Frame(self.window, padding=12)
        buttons.pack(fill=X)
        ttk.Button(buttons, text="Select all", command=lambda: self._set_all(True)).pack(side="left")
        ttk.Button(buttons, text="Clear", command=lambda: self._set_all(False)).pack(side="left", padx=8)
        ttk.Button(buttons, text="Continue", command=self.accept).pack(side="right")
        ttk.Button(buttons, text="Cancel", command=self.cancel).pack(side="right", padx=8)
        self.window.protocol("WM_DELETE_WINDOW", self.cancel)

    def _set_all(self, value: bool) -> None:
        self.date_list.selection_clear(0, tk.END)
        if value and self.dates:
            self.date_list.selection_set(0, tk.END)

    def accept(self) -> None:
        selected_indices = self.date_list.curselection()
        self.selected = [self.dates[index] for index in selected_indices]
        if not self.selected:
            messagebox.showerror("Date selection", "Select at least one date.", parent=self.window)
            return
        self.window.destroy()

    def cancel(self) -> None:
        self.selected = None
        self.window.destroy()


class SpeedApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Wavetec speed data to Excel")
        self.root.geometry("900x205")
        self.input_paths = tk.StringVar()
        self.output_path = tk.StringVar()
        self.convert_button: ttk.Button
        self._build_ui()

    def _build_ui(self) -> None:
        controls = ttk.Frame(self.root, padding=16)
        controls.pack(fill=X)
        ttk.Label(controls, text="Input ZIP/TXT file(s):").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(controls, textvariable=self.input_paths).grid(row=0, column=1, sticky="ew", pady=6)
        ttk.Button(controls, text="Browse...", command=self.choose_input).grid(row=0, column=2, padx=(8, 0), pady=6)
        ttk.Label(controls, text="Excel output file:").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(controls, textvariable=self.output_path).grid(row=1, column=1, sticky="ew", pady=6)
        ttk.Button(controls, text="Browse...", command=self.choose_output).grid(row=1, column=2, padx=(8, 0), pady=6)
        self.convert_button = ttk.Button(controls, text="Convert (this may take a long time)", command=self.convert)
        self.convert_button.grid(row=0, column=3, rowspan=2, padx=(14, 0), ipadx=18)
        controls.columnconfigure(1, weight=1)
        ttk.Label(self.root, text="Select multiple files in the file dialog with Ctrl or Shift, then click Open.", padding=(16, 0)).pack(anchor="w")

    def choose_input(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Select ZIP or TXT files (use Ctrl/Shift for multiple files)",
            filetypes=[("ZIP and TXT files", "*.zip *.txt"), ("ZIP files", "*.zip"), ("TXT files", "*.txt"), ("All files", "*.*")],
        )
        if paths:
            self.input_paths.set(";".join(paths))

    def choose_output(self) -> None:
        path = filedialog.asksaveasfilename(title="Save Excel workbook", defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")])
        if path:
            self.output_path.set(path)

    def convert(self) -> None:
        paths = [path for path in self.input_paths.get().split(";") if path]
        if not paths:
            messagebox.showerror("Conversion error", "Select at least one input ZIP or TXT file.")
            return
        output_path = self.output_path.get().strip()
        if not output_path:
            messagebox.showerror("Conversion error", "Select an Excel output file.")
            return
        self.convert_button.configure(state="disabled")
        progress = ProgressDialog(self.root)
        worker = threading.Thread(target=self._parse_worker, args=(paths, output_path, progress), daemon=True)
        worker.start()

    def _parse_worker(self, paths: list[str], output_path: str, progress: ProgressDialog) -> None:
        try:
            measurements, _skipped = load_archives(paths, progress=lambda current, total, name: self.root.after(0, progress.update, 5 + int(current / total * 40), f"Parsed {name}"))
            if not measurements:
                raise ValueError("No complete pairs of measurements were found in the selected ZIP/TXT files.")
            self.root.after(0, self._show_date_selection, measurements, output_path, progress)
        except Exception as error:
            self.root.after(0, self._conversion_error, progress, error, traceback.format_exc())

    def _show_date_selection(self, measurements: list[Measurement], output_path: str, progress: ProgressDialog) -> None:
        progress.window.grab_release()
        progress.window.destroy()
        dialog = DateSelectionDialog(self.root, sorted({item.timestamp.date() for item in measurements}))
        self.root.wait_window(dialog.window)
        if dialog.selected is None:
            self.convert_button.configure(state="normal")
            return
        selected_dates = set(dialog.selected)
        selected_measurements = [item for item in measurements if item.timestamp.date() in selected_dates]
        export_progress = ProgressDialog(self.root)
        threading.Thread(target=self._export_worker, args=(selected_measurements, output_path, export_progress), daemon=True).start()

    def _export_worker(self, measurements: list[Measurement], output_path: str, progress: ProgressDialog) -> None:
        try:
            export_workbook(measurements, output_path, progress=lambda percent, message: self.root.after(0, progress.update, 40 + int(percent * 0.6), message))
            self.root.after(0, self._conversion_success, progress, output_path)
        except Exception as error:
            self.root.after(0, self._conversion_error, progress, error, traceback.format_exc())

    def _conversion_success(self, progress: ProgressDialog, output_path: str) -> None:
        progress.update(100, "Finished")
        progress.window.grab_release()
        progress.window.destroy()
        self.convert_button.configure(state="normal")
        messagebox.showinfo("Conversion complete", f"Excel workbook created:\n{output_path}")

    def _conversion_error(self, progress: ProgressDialog, error: Exception, details: str) -> None:
        progress.window.grab_release()
        progress.window.destroy()
        self.convert_button.configure(state="normal")
        messagebox.showerror("Conversion error", f"{type(error).__name__}: {error}\n\n{details}")


def main() -> None:
    root = tk.Tk()
    try:
        from tkinter import font
        root.option_add("*Font", ("Segoe UI", 9))
        font.nametofont("TkDefaultFont").configure(size=9)
    except tk.TclError:
        pass
    SpeedApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()